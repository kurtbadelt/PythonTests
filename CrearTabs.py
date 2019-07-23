# coding=utf-8
from openpyxl import load_workbook

class CrearTabs():
    def ejecutarCrearTabs():

        print("El primer paso es crear las pestañas")
        inputVar = input("Da enter para iniciar el proceso-> ")
        respuesta = input("La quincena tiene 2 o 3 semanas? ")
        if int(respuesta) == 2:
            print("-Quincena de dos semanas-")
            print("Creando pestañas por comedor....")
            wb = load_workbook('ControlAsistenciaProcomin_Template.xlsx')
            wb.create_sheet('ADO')
            wb.create_sheet('PRICE SHOES')
            wb.create_sheet('MABESA')
            wb.create_sheet('METALOIDES')
            wb.create_sheet('ZUUM')
            wb.create_sheet('PODER JUDICIAL')
            wb.create_sheet('SOLUTIA')
            wb.create_sheet('SUMITOMO TLAXCALA')
            wb.create_sheet('ADIENT TLAXCALA')
            wb.create_sheet('GESTAMP FINSA')
            wb.create_sheet('ADIENT FINSA')
            wb.create_sheet('BROSE')
            wb.create_sheet('KAYSER')
            wb.create_sheet('SMP FINSA')
            wb.create_sheet('ROA')
            wb.create_sheet('PISTONES')
            wb.create_sheet('GESTAMP 2')
            wb.create_sheet('THYSSEN 1')
            wb.create_sheet('THYSSEN 2')
            wb.create_sheet('MAQUINSA')
            wb.create_sheet('POSCO')
            wb.create_sheet('BODEGA')
            wb.create_sheet('OFICINAS')
            wb.create_sheet('SUMITOMO ATLIXCO')
            wb.create_sheet('FAURECIA')
            wb.create_sheet('COJINETES')
            wb.create_sheet('AUDI')
            wb.create_sheet('SMP TLAXCALA')
            wb.create_sheet('PODER J. TLAXCALA')
            wb.create_sheet('INTEVA')
            wb.create_sheet('PAKAR SHOES')
            wb.save('ControlAsistenciaProcomin_V1.xlsx')
        else:
            print("-Quincena de tres semanas-")
            print("Creando pestañas por comedor....")
            wb = load_workbook('ControlAsistenciaProcomin_Template2.xlsx')
            wb.create_sheet('ADO')
            wb.create_sheet('PRICE SHOES')
            wb.create_sheet('MABESA')
            wb.create_sheet('METALOIDES')
            wb.create_sheet('ZUUM')
            wb.create_sheet('PODER JUDICIAL')
            wb.create_sheet('SOLUTIA')
            wb.create_sheet('SUMITOMO TLAXCALA')
            wb.create_sheet('ADIENT TLAXCALA')
            wb.create_sheet('GESTAMP FINSA')
            wb.create_sheet('ADIENT FINSA')
            wb.create_sheet('BROSE')
            wb.create_sheet('KAYSER')
            wb.create_sheet('SMP FINSA')
            wb.create_sheet('ROA')
            wb.create_sheet('PISTONES')
            wb.create_sheet('GESTAMP 2')
            wb.create_sheet('THYSSEN 1')
            wb.create_sheet('THYSSEN 2')
            wb.create_sheet('MAQUINSA')
            wb.create_sheet('POSCO')
            wb.create_sheet('BODEGA')
            wb.create_sheet('OFICINAS')
            wb.create_sheet('SUMITOMO ATLIXCO')
            wb.create_sheet('FAURECIA')
            wb.create_sheet('COJINETES')
            wb.create_sheet('AUDI')
            wb.create_sheet('SMP TLAXCALA')
            wb.create_sheet('PODER J. TLAXCALA')
            wb.create_sheet('INTEVA')
            wb.create_sheet('PAKAR SHOES')
            wb.save('ControlAsistenciaProcomin_V1.xlsx')





print("\t***************************************************************")
print("\t***  Bienvenido, vamos a crear los archivos para la nómina  ***")
print("\t***************************************************************")

CrearTabs.ejecutarCrearTabs()

