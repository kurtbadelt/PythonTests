# coding=utf-8
from openpyxl import load_workbook
from datetime import datetime


# Definición de clases
class DatosEmpleado:
    def __init__(self, idEmpleado, nombreEmpleado, codigoComedor, comedor,row):
        self.idEmpleado = idEmpleado
        self.nombreEmpleado = nombreEmpleado
        self.codigoComedor = codigoComedor
        self.comedor = comedor
        self.row = row


    def to_string(self):
        print(self.nombreEmpleado)

class Asistencia:
    def __init__(self,idEmpleado,nombreEmpleado,tiempo,estado,terminal):
        self.idEmpleado = idEmpleado
        self.nombreEmpleado = nombreEmpleado
        self.tiempo = tiempo
        self.estado = estado
        self.terminal = terminal
        self.datetime_object = datetime.strptime(self.tiempo, '%m/%d/%Y %H:%M:%S')



# Leer el archivo del control de asistencias
def cargarArchivosEntrada(reporteZKTime):
    return  load_workbook(reporteZKTime)


wb = load_workbook('ControlAsistenciaProcomin_V3.xlsx')
# Objetos hojas del control de asistencias
sheetDatosEmpleados = wb['Datos Empleados']
sheetADO = wb['ADO']
sheetPriceShoes = wb['PRICE SHOES']
sheetMabesa = wb['MABESA']
sheetMetaloides = wb['METALOIDES']
sheetZuum = wb['ZUUM']
sheetPoderJudicial = wb['PODER JUDICIAL']
sheetSolutia = wb['SOLUTIA']
sheetSumitomoTlaxcala = wb['SUMITOMO TLAXCALA']
sheetAdientTlaxcala = wb['ADIENT TLAXCALA']
sheetGestampFinsa = wb['GESTAMP FINSA']
sheetAdientFinsa = wb['ADIENT FINSA']
sheetBrose = wb['BROSE']
sheetKayser = wb['KAYSER']
sheetSmpFinsa = wb['SMP FINSA']
sheetROA = wb['ROA']
sheetPistones = wb['PISTONES']
sheetGestamp2 = wb['GESTAMP 2']
sheetThyssen1 = wb['THYSSEN 1']
sheetThyssen2 = wb['THYSSEN 2']
sheetMaquinsa = wb['MAQUINSA']
sheetPosco = wb['POSCO']
sheetBodega = wb['BODEGA']
sheetOficina = wb['OFICINAS']
sheetSumitomoAtlixco = wb['SUMITOMO ATLIXCO']
sheetFaurecia = wb['FAURECIA']
sheetCojinetes = wb['COJINETES']
sheetAudi = wb['AUDI']
sheetSmpTlaxcala = wb['SMP TLAXCALA']
sheetPjTlaxcala = wb['PODER J. TLAXCALA']
sheetInteva = wb['INTEVA']
sheetPakar = wb['PAKAR SHOES']





#Lista de Empleados de control asistencias
listaAsistencias = []



# Recorrer la hoja de Datos Empleado y guardarlos en un arreglo
def leerDatosAsistencia(reporteZKTime):

    wbReporte = cargarArchivosEntrada(reporteZKTime)
    # Objeto hoja del reporte de zktime
    sheetReporte = wbReporte['Hoja1']
    # Contador de lineas en hoja datos empleados
    numeroIncidencias = sheetReporte.max_row
    print("Leyendo Datos Asistencia")
    for i in range(2, numeroIncidencias + 1):
        idEmpleado = int(sheetReporte.cell(row=i, column=1).value)
        nombreEmpleado = sheetReporte.cell(row=i, column=2).value
        tiempo = sheetReporte.cell(row=i, column=3).value
        estado = sheetReporte.cell(row=i, column=4).value
        terminal = sheetReporte.cell(row=i, column=5).value
        asistencia = Asistencia(idEmpleado,nombreEmpleado,tiempo,estado,terminal)
        listaAsistencias.append(asistencia)


#lee la fecha de inicio de una hoja
def leerFechaInicio(objetoHoja):
    datetime_object = datetime.strptime(objetoHoja.cell(row=2,column=2).value, '%m/%d/%Y %H:%M:%S')
    return datetime_object

#lee la fecha final de una hoja
def leerFechaFin(objetoHoja):
    datetime_object = datetime.strptime(objetoHoja.cell(row=3,column=2).value, '%m/%d/%Y %H:%M:%S')
    return datetime_object

#lee los empleados de una hoja
def leerEmpleados(objetoHoja):
    listaEmpleados = []
    numeroEmpleados = objetoHoja.max_row-8
    for i in range(8,numeroEmpleados+1):
        idEmpleado = objetoHoja.cell(row=i,column=1).value
        nombreEmpleado = ""
        codigoComedor = ""
        comedor=""
        if not idEmpleado == None:
            empleado = DatosEmpleado(idEmpleado, nombreEmpleado, codigoComedor, comedor, i)
            listaEmpleados.append(empleado)

    return listaEmpleados

#valida si una fecha se encuentra dentro del rango establecido
#primer dato es la fecha a comparar, segundo y tercero el rango
def isFechaInRango(fecha,fechaInicio,fechaFin):
    if fechaInicio <= fecha <= fechaFin:

        return True
    else:

        return False

def contarEntradas(numeroEmpleado,fechaInicio,fechaFin):
    contador=0
    for i in range(len(listaAsistencias)):
        if numeroEmpleado == listaAsistencias[i].idEmpleado:
            if (listaAsistencias[i].estado == "Entrada") or (listaAsistencias[i].estado == "Entrada T.E."):
                if isFechaInRango(listaAsistencias[i].datetime_object,fechaInicio,fechaFin):
                    contador+=1

    return contador

def seleccionarColumnaAsistencia(renglon,objetoHoja):

    valor = objetoHoja.cell(row=renglon,column=22).value
    if valor == "A":

        return 28
    else:

        return 22

def calcularAsistencia(objetoHoja):
    fechaInicio = leerFechaInicio(objetoHoja)
    fechaFin = leerFechaFin(objetoHoja)
    listaEmpleados = leerEmpleados(objetoHoja)
    numeroEmpleados = len(listaEmpleados)

    for i in range(0,numeroEmpleados):
        asistencia = contarEntradas(listaEmpleados[i].idEmpleado,fechaInicio,fechaFin)
        objetoHoja.cell(row=listaEmpleados[i].row,column=seleccionarColumnaAsistencia(listaEmpleados[i].row,objetoHoja)).value = asistencia

def esNumero(id):

    if id is not None:
        try:
            val = int(id)

        except ValueError:

            return False
        
        return True


def limpiarHoja(objetoHoja):
    colCount = objetoHoja.max_column +1
    for i in range(8,101):
        id = objetoHoja.cell(row=i,column=1).value
        if not esNumero(id):
            for j in range(1, colCount):
                objetoHoja.cell(row=i, column=j).value = ""





#"general_18-24-junio.xlsx
def ejecutarCalcularAsistencia(reporteZKTime,archivoSalida):
    print("Calculando la asistencia")
    leerDatosAsistencia(reporteZKTime)
    calcularAsistencia(sheetADO)
    limpiarHoja(sheetADO)
    calcularAsistencia(sheetPriceShoes)
    limpiarHoja(sheetPriceShoes)
    calcularAsistencia(sheetMabesa)
    limpiarHoja(sheetMabesa)
    calcularAsistencia(sheetMetaloides)
    limpiarHoja(sheetMetaloides)
    calcularAsistencia(sheetZuum)
    limpiarHoja(sheetZuum)
    calcularAsistencia(sheetPoderJudicial)
    limpiarHoja(sheetPoderJudicial)
    calcularAsistencia(sheetSolutia)
    limpiarHoja(sheetSolutia)
    calcularAsistencia(sheetSumitomoTlaxcala)
    limpiarHoja(sheetSumitomoTlaxcala)
    calcularAsistencia(sheetAdientTlaxcala)
    limpiarHoja(sheetAdientTlaxcala)
    calcularAsistencia(sheetGestampFinsa)
    limpiarHoja(sheetGestampFinsa)
    calcularAsistencia(sheetAdientFinsa)
    limpiarHoja(sheetAdientFinsa)
    calcularAsistencia(sheetBrose)
    limpiarHoja(sheetBrose)
    calcularAsistencia(sheetKayser)
    limpiarHoja(sheetKayser)
    calcularAsistencia(sheetSmpFinsa)
    limpiarHoja(sheetSmpFinsa)
    calcularAsistencia(sheetROA)
    limpiarHoja(sheetROA)
    calcularAsistencia(sheetPistones)
    limpiarHoja(sheetPistones)
    calcularAsistencia(sheetGestamp2)
    limpiarHoja(sheetGestamp2)
    calcularAsistencia(sheetThyssen1)
    limpiarHoja(sheetThyssen1)
    calcularAsistencia(sheetThyssen2)
    limpiarHoja(sheetThyssen2)
    calcularAsistencia(sheetMaquinsa)
    limpiarHoja(sheetMaquinsa)
    calcularAsistencia(sheetPosco)
    limpiarHoja(sheetPosco)
    calcularAsistencia(sheetBodega)
    limpiarHoja(sheetBodega)
    calcularAsistencia(sheetOficina)
    limpiarHoja(sheetOficina)
    calcularAsistencia(sheetSumitomoAtlixco)
    limpiarHoja(sheetSumitomoAtlixco)
    calcularAsistencia(sheetFaurecia)
    limpiarHoja(sheetFaurecia)
    calcularAsistencia(sheetCojinetes)
    limpiarHoja(sheetCojinetes)
    calcularAsistencia(sheetAudi)
    limpiarHoja(sheetAudi)
    calcularAsistencia(sheetSmpTlaxcala)
    limpiarHoja(sheetSmpTlaxcala)
    calcularAsistencia(sheetPjTlaxcala)
    limpiarHoja(sheetPjTlaxcala)
    calcularAsistencia(sheetInteva)
    limpiarHoja(sheetInteva)
    calcularAsistencia(sheetPakar)
    limpiarHoja(sheetPakar)
    wb.save(archivoSalida)  # guardar un archivo nuevo

#método que valida que los archivos existen o no.
def validarEntradas(archivo1):
    try:
        wb = load_workbook(archivo1)
    except IOError:
        print("No pudimos encontrar el archivo")
        print(archivo1)
        return False
    return True


reporteZKTime = input("Ingresa el nombre del archivo del sistema de asistencias, ejemplo: general_18-24-junio.xlsx -> ")
archivoSalida = input("¿Cómo quieres nombrar a tu archivo?, Ejemplo: NominaPrimeraQuincenaMayo2018.xlsx -> ")

if validarEntradas(reporteZKTime):
    ejecutarCalcularAsistencia(reporteZKTime, archivoSalida)
    print("Proceso terminado, puedes abrir el archivo: ")
    print(archivoSalida)

