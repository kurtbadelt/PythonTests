# coding=utf-8
from __future__ import absolute_import
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet import Worksheet
from copy import copy
from datetime import datetime
import sys
# Copyright (c) 2010-2017 openpyxl

class WorksheetCopy(object):
    """
    Copy the values, styles, dimensions and merged cells from one worksheet
    to another within the same workbook.
    """

    def __init__(self, source_worksheet, target_worksheet):
        self.source = source_worksheet
        self.target = target_worksheet
        self._verify_resources()


    def _verify_resources(self):

        if (not isinstance(self.source, Worksheet)
            and not isinstance(self.target, Worksheet)):
            raise TypeError("Can only copy worksheets")

        if self.source is self.target:
            raise ValueError("Cannot copy a worksheet to itself")

        if self.source.parent != self.target.parent:
            raise ValueError('Cannot copy between worksheets from different workbooks')


    def copy_worksheet(self):
        self._copy_cells()
        self._copy_dimensions()

        self.target.sheet_format = copy(self.source.sheet_format)
        self.target.sheet_properties = copy(self.source.sheet_properties)
        #no funciona copiar merged cells
       # self.target.merged_cells = copy(self.source.merged_cells)



    def _copy_cells(self):
        for (row, col), source_cell  in self.source._cells.items():
            target_cell = self.target.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


    def _copy_dimensions(self):
        for attr in ('row_dimensions', 'column_dimensions'):
            src = getattr(self.source, attr)
            target = getattr(self.target, attr)
            for key, dim in src.items():
                target[key] = copy(dim)
                target[key].worksheet = self.target

#lee la fecha de inicio de una hoja
def escribirFechaInicio(objetoHoja,fechaInicio):
    #"06/15/2018 15:45:35"
    objetoHoja.cell(row=2, column=2).value =fechaInicio



#lee la fecha final de una hoja
def escribirFechaFin(objetoHoja,fechaFin):
    #
    objetoHoja.cell(row=3, column=2).value = fechaFin


#Zona de ejecución
def ejecutarCopiarContenido(fechaInicio,fechaFin):
    print("Ejecutando CopiarContenido....")
    wb = load_workbook('ControlAsistenciaProcomin_V1.xlsx')
    sheetBase = wb['HojaBase']
    sheetADO = wb['ADO']
    sheetPriceShoes = wb['PRICE SHOES']
    sheetMabesa = wb['MABESA']
    sheetMetaloides = wb['METALOIDES']
    sheetZuum = wb['ZUUM']
    sheetPoderJudicial = wb['PODER JUDICIAL']
    sheetSolutia = wb['SOLUTIA']
    sheetSumitomoTlaxcala = wb['SUMITOMO TLAXCALA']
    sheetAdientTlaxcala = wb['ADIENT TLAXCALA']
    sheetGestampFinsa = wb['GESTAMP FINSA']
    sheetAdientFinsa = wb['ADIENT FINSA']
    sheetBrose = wb['BROSE']
    sheetKayser = wb['KAYSER']
    sheetSmpFinsa = wb['SMP FINSA']
    sheetROA = wb['ROA']
    sheetPistones = wb['PISTONES']
    sheetGestamp2 = wb['GESTAMP 2']
    sheetThyssen1 = wb['THYSSEN 1']
    sheetThyssen2 = wb['THYSSEN 2']
    sheetMaquinsa = wb['MAQUINSA']
    sheetPosco = wb['POSCO']
    sheetBodega = wb['BODEGA']
    sheetOficina = wb['OFICINAS']
    sheetSumitomoAtlixco = wb['SUMITOMO ATLIXCO']
    sheetFaurecia = wb['FAURECIA']
    sheetCojinetes = wb['COJINETES']
    sheetAudi = wb['AUDI']
    sheetSmpTlaxcala = wb['SMP TLAXCALA']
    sheetPjTlaxcala = wb['PODER J. TLAXCALA']
    sheetInteva = wb['INTEVA']
    sheetPakar = wb['PAKAR SHOES']

    xeroxMachine = WorksheetCopy(sheetBase, sheetADO)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetADO,fechaFin)
    escribirFechaInicio(sheetADO,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPriceShoes)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPriceShoes,fechaFin)
    escribirFechaInicio(sheetPriceShoes,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPoderJudicial)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPoderJudicial,fechaFin)
    escribirFechaInicio(sheetPoderJudicial,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetMabesa)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetMabesa,fechaFin)
    escribirFechaInicio(sheetMabesa,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetMetaloides)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetMetaloides,fechaFin)
    escribirFechaInicio(sheetMetaloides,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetZuum)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetZuum,fechaFin)
    escribirFechaInicio(sheetZuum,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetSolutia)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetSolutia,fechaFin)
    escribirFechaInicio(sheetSolutia,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetSumitomoTlaxcala)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetSumitomoTlaxcala,fechaFin)
    escribirFechaInicio(sheetSumitomoTlaxcala,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetAdientTlaxcala)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetAdientTlaxcala,fechaFin)
    escribirFechaInicio(sheetAdientTlaxcala,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetGestampFinsa)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetGestampFinsa,fechaFin)
    escribirFechaInicio(sheetGestampFinsa,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetAdientFinsa)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetAdientFinsa,fechaFin)
    escribirFechaInicio(sheetAdientFinsa,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetBrose)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetBrose,fechaFin)
    escribirFechaInicio(sheetBrose,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetKayser)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetKayser,fechaFin)
    escribirFechaInicio(sheetKayser,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetSmpFinsa)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetSmpFinsa,fechaFin)
    escribirFechaInicio(sheetSmpFinsa,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetROA)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetROA,fechaFin)
    escribirFechaInicio(sheetROA,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPistones)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPistones,fechaFin)
    escribirFechaInicio(sheetPistones,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetGestamp2)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetGestamp2,fechaFin)
    escribirFechaInicio(sheetGestamp2,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetThyssen1)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetThyssen1,fechaFin)
    escribirFechaInicio(sheetThyssen1,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetThyssen2)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetThyssen2,fechaFin)
    escribirFechaInicio(sheetThyssen2,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetMaquinsa)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetMaquinsa,fechaFin)
    escribirFechaInicio(sheetMaquinsa,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPosco)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPosco,fechaFin)
    escribirFechaInicio(sheetPosco,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetBodega)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetBodega,fechaFin)
    escribirFechaInicio(sheetBodega,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetOficina)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetOficina,fechaFin)
    escribirFechaInicio(sheetOficina,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetSumitomoAtlixco)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetSumitomoAtlixco,fechaFin)
    escribirFechaInicio(sheetSumitomoAtlixco,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetFaurecia)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetFaurecia,fechaFin)
    escribirFechaInicio(sheetFaurecia,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetCojinetes)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetCojinetes,fechaFin)
    escribirFechaInicio(sheetCojinetes,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetAudi)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetAudi,fechaFin)
    escribirFechaInicio(sheetAudi,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetSmpTlaxcala)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetSmpTlaxcala,fechaFin)
    escribirFechaInicio(sheetSmpTlaxcala,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPjTlaxcala)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPjTlaxcala,fechaFin)
    escribirFechaInicio(sheetPjTlaxcala,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetInteva)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetInteva,fechaFin)
    escribirFechaInicio(sheetInteva,fechaInicio)

    xeroxMachine = WorksheetCopy(sheetBase, sheetPakar)
    xeroxMachine.copy_worksheet()
    escribirFechaFin(sheetPakar,fechaFin)
    escribirFechaInicio(sheetPakar,fechaInicio)

    renglonCodigo = 4
    columnaCodigo = 2
    sheetADO.cell(row=renglonCodigo, column=columnaCodigo).value = 50
    sheetPriceShoes.cell(row=renglonCodigo, column=columnaCodigo).value = 51
    sheetMabesa.cell(row=renglonCodigo, column=columnaCodigo).value = 52
    sheetMetaloides.cell(row=renglonCodigo, column=columnaCodigo).value = 53
    sheetZuum.cell(row=renglonCodigo, column=columnaCodigo).value = 54
    sheetPoderJudicial.cell(row=renglonCodigo, column=columnaCodigo).value = 55
    sheetSolutia.cell(row=renglonCodigo, column=columnaCodigo).value = 58
    sheetSumitomoTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value = 59
    sheetAdientTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value = 60
    sheetGestampFinsa.cell(row=renglonCodigo, column=columnaCodigo).value = 61
    sheetGestampFinsa.cell(row=renglonCodigo, column=columnaCodigo).value = 62
    sheetBrose.cell(row=renglonCodigo, column=columnaCodigo).value = 63
    sheetKayser.cell(row=renglonCodigo, column=columnaCodigo).value = 64
    sheetSmpFinsa.cell(row=renglonCodigo, column=columnaCodigo).value = 65
    sheetROA.cell(row=renglonCodigo, column=columnaCodigo).value = 66
    sheetPistones.cell(row=renglonCodigo, column=columnaCodigo).value = 67
    sheetGestamp2.cell(row=renglonCodigo, column=columnaCodigo).value = 68
    sheetThyssen1.cell(row=renglonCodigo, column=columnaCodigo).value = 69
    sheetThyssen2.cell(row=renglonCodigo, column=columnaCodigo).value = 70
    sheetMaquinsa.cell(row=renglonCodigo, column=columnaCodigo).value = 71
    sheetPosco.cell(row=renglonCodigo, column=columnaCodigo).value = 72
    sheetBodega.cell(row=renglonCodigo, column=columnaCodigo).value = 73
    sheetOficina.cell(row=renglonCodigo, column=columnaCodigo).value = 74
    sheetSumitomoAtlixco.cell(row=renglonCodigo, column=columnaCodigo).value = 75
    sheetFaurecia.cell(row=renglonCodigo, column=columnaCodigo).value = 76
    sheetCojinetes.cell(row=renglonCodigo, column=columnaCodigo).value = 77
    sheetAudi.cell(row=renglonCodigo, column=columnaCodigo).value = 78
    sheetSmpTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value = 80
    sheetPjTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value = 81
    sheetInteva.cell(row=renglonCodigo, column=columnaCodigo).value = 82
    sheetPakar.cell(row=renglonCodigo, column=columnaCodigo).value = 83

    wb.save('ControlAsistenciaProcomin_V2.xlsx')  # guardar un archivo nuevo

#Metodo que valida si las entradas tienene l formato de fecha correcto
def validarEntradas(fechaI,fechaF):
    try:
        datetime_objectI = datetime.strptime(fechaI, '%m/%d/%Y %H:%M:%S')
        datetime_objectF = datetime.strptime(fechaF, '%m/%d/%Y %H:%M:%S')
    except ValueError:
        print("las fechas no tienen el formato correcto")
        return False
    return True

    return True

print("Ahora vamos a llenar las pestañas con la información y formulas")
print("Ingresa la fecha de inicio de la quincena")
fechaInicio_Input = input("Fecha de inicio en formato de la fecha: mm/dd/aaaa, ejemplo: 06/15/2018 -> ")
fechaInicio_Input+=" 00:00:00"
print("Primer Día de la quincena:")
print(fechaInicio_Input)
print("Ingresa la fecha en que termina la quincena")
fechaFin_Input = input("Fecha de inicio en formato de la fecha: mm/dd/aaaa, ejemplo: 06/30/2018 -> ")
fechaFin_Input+=" 00:00:00"
print("Último día de la quincena")
print(fechaFin_Input)

if validarEntradas(fechaInicio_Input,fechaFin_Input):
    ejecutarCopiarContenido(fechaInicio_Input, fechaFin_Input)
else:
    print("se cerro el programa por error en las fechas")
    sys.exit()



