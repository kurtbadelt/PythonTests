# coding=utf-8
from openpyxl import load_workbook

# Definición de clases
class DatosEmpleado:
    def __init__(self, idEmpleado, nombreEmpleado, codigoComedor, comedor,row):
        self.idEmpleado = idEmpleado
        self.nombreEmpleado = nombreEmpleado
        self.codigoComedor = codigoComedor
        self.comedor = comedor
        self.row = row


    def to_string(self):
        print(self.nombreEmpleado)

#Workbooks
wb = load_workbook('ControlAsistenciaProcomin_V2.xlsx')


#Worksheets
sheetDatosEmpleados = wb['Datos Empleados']



#Contador de lineas en hoja datos empleados
numeroEmpleados = sheetDatosEmpleados.max_row

#Lista de Empleados de control asistencias
listaEmpleados = []

#lista de Empleados de empleadosProcomin
listaEmpleadosReporteZK = []


def leerDatosEmpleadosProcomin(nombreReporte):
    wbEmpleados = load_workbook(nombreReporte)
    sheetEmpleadosReporteZK = wbEmpleados['empleados_procomin']
    numeroEmpleadosProcomin = sheetEmpleadosReporteZK.max_row
    print("Leyendo Datos Empleados Procomin")
    for i in range(2,numeroEmpleadosProcomin + 1):
        idEmpleado = sheetEmpleadosReporteZK.cell(row=i,column=1).value
        nombreEmpleado = sheetEmpleadosReporteZK.cell(row=i,column=2).value
        codigoComedor = sheetEmpleadosReporteZK.cell(row=i,column=3).value
        nombreComedor = sheetEmpleadosReporteZK.cell(row=i,column=4).value
        empleado=DatosEmpleado(idEmpleado,nombreEmpleado,codigoComedor,nombreComedor,i)
        listaEmpleadosReporteZK.append(empleado)

# Recorrer la hoja de Datos Empleado y guardarlos en un arreglo
def leerDatosEmpleados():
    print("Leyendo Datos Empleados")
    for i in range(2, numeroEmpleados + 1):
        idEmpleado = sheetDatosEmpleados.cell(row=i, column=1).value
        nombreEmpleado = sheetDatosEmpleados.cell(row=i, column=2).value
        codigoComedor = sheetDatosEmpleados.cell(row=i, column=3).value
        nombreComedor = sheetDatosEmpleados.cell(row=i, column=4).value
        empleado = DatosEmpleado(idEmpleado, nombreEmpleado, codigoComedor, nombreComedor,i)
        listaEmpleados.append(empleado)


def containsID(lista1,numero):
    lenLista1 = len(lista1)
    contador =0;

    for i in range(lenLista1):
        if lista1[i].idEmpleado == numero:
            contador += 1

    if contador > 0:
        return True
    else:
        return False



def containsNombre(lista1,nombre):
    listaNombres = []
    for i in range(len(lista1)):
        listaNombres.append(lista1[i].nombreEmpleado)
    if nombre in str(listaNombres):
        return True
    else:
        return False


def getRow(lista1, numero):
    lenLista1 = len(lista1)
    for i in range(lenLista1):
        if lista1[i].idEmpleado == numero:

            return i+2

    # Abre dos archivos de Excel (Reporte de empleados en ZKTime y Control de nóminas
    # Busca los empleados que no están en el control de nóminas y los agrega
    # Busca los cambios en los nombres y los actualiza // Esto aun no funciona bien.
def actualizarEmpleados(nombreReporte):
        leerDatosEmpleados()
        leerDatosEmpleadosProcomin(nombreReporte)
        lenListaProcomin = len(listaEmpleadosReporteZK)

        for i in range(lenListaProcomin):
            if not containsID(listaEmpleados, listaEmpleadosReporteZK[i].idEmpleado):
                print("No encontrado")
                print(listaEmpleadosReporteZK[i].idEmpleado)
                nextRow = sheetDatosEmpleados.max_row + 1
                sheetDatosEmpleados.cell(row=nextRow, column=1).value = listaEmpleadosReporteZK[i].idEmpleado
                sheetDatosEmpleados.cell(row=nextRow, column=2).value = listaEmpleadosReporteZK[i].nombreEmpleado
                sheetDatosEmpleados.cell(row=nextRow, column=3).value = listaEmpleadosReporteZK[i].codigoComedor
                sheetDatosEmpleados.cell(row=nextRow, column=4).value = listaEmpleadosReporteZK[i].comedor

            else:
                if not containsNombre(listaEmpleados, listaEmpleadosReporteZK[i].nombreEmpleado):
                    print("Nombre No Encotrado")
                    print(listaEmpleadosReporteZK[i].nombreEmpleado,listaEmpleadosReporteZK[i].idEmpleado)
                    rowToUpdate = getRow(listaEmpleados,listaEmpleadosReporteZK[i].idEmpleado)
                    print("Actualizando Nombre....")
                    sheetDatosEmpleados.cell(row=rowToUpdate, column=2).value = listaEmpleadosReporteZK[i].nombreEmpleado

# empleados_procomin_prueba.xlsx
def ejecutarActualizarEmpleados(nombreReporte):
    print("Vamos a validar que la base de datos de empleados esté actualizada")
    print("Actualizando Empleados")
    actualizarEmpleados(nombreReporte)
    wb.save('ControlAsistenciaProcomin_V2.xlsx')

#método que valida si el archivo existe o no.
def validaEntrada(nombreReporte):
    try:
        wb = load_workbook(nombreReporte)
    except IOError:
        print("No pudimos encontrar el archivo")
        print(nombreReporte)
        return False
    return True

nombreReporte = input("Ingresa el nombre del reporte donde están los empleados, Ejemplo: empleados_procomin_prueba.xlsx -> ")

if validaEntrada(nombreReporte):
    ejecutarActualizarEmpleados(nombreReporte)

