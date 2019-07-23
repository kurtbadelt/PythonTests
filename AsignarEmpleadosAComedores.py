# coding=utf-8
from openpyxl import load_workbook


# Definición de clases
class DatosEmpleado:
    def __init__(self, idEmpleado, nombreEmpleado, codigoComedor, comedor):
        self.idEmpleado = idEmpleado
        self.nombreEmpleado = nombreEmpleado
        self.codigoComedor = codigoComedor
        self.comedor = comedor

    def to_string(self):
        print(self.nombreEmpleado)



# Leer el archivo del control de asistencias
wb = load_workbook('ControlAsistenciaProcomin_V2.xlsx')


# Objetos hojas del control de asistencias
# Sería mejor tener un arreglo de objetos tipo Sheet, para recorrerlo
sheetDatosEmpleados = wb['Datos Empleados']
sheetADO = wb['ADO']
sheetPriceShoes = wb['PRICE SHOES']
sheetMabesa = wb['MABESA']
sheetMetaloides = wb['METALOIDES']
sheetZuum = wb['ZUUM']
sheetPoderJudicial = wb['PODER JUDICIAL']
sheetSolutia = wb['SOLUTIA']
sheetSumitomoTlaxcala = wb['SUMITOMO TLAXCALA']
sheetAdientTlaxcala = wb['ADIENT TLAXCALA']
sheetGestampFinsa = wb['GESTAMP FINSA']
sheetAdientFinsa = wb['ADIENT FINSA']
sheetBrose = wb['BROSE']
sheetKayser = wb['KAYSER']
sheetSmpFinsa = wb['SMP FINSA']
sheetROA = wb['ROA']
sheetPistones = wb['PISTONES']
sheetGestamp2 = wb['GESTAMP 2']
sheetThyssen1 = wb['THYSSEN 1']
sheetThyssen2 = wb['THYSSEN 2']
sheetMaquinsa = wb['MAQUINSA']
sheetPosco = wb['POSCO']
sheetBodega = wb['BODEGA']
sheetOficina = wb['OFICINAS']
sheetSumitomoAtlixco = wb['SUMITOMO ATLIXCO']
sheetFaurecia = wb['FAURECIA']
sheetCojinetes = wb['COJINETES']
sheetAudi = wb['AUDI']
sheetSmpTlaxcala = wb['SMP TLAXCALA']
sheetPjTlaxcala = wb['PODER J. TLAXCALA']
sheetInteva = wb['INTEVA']
sheetPakar = wb['PAKAR SHOES']



#Objetos donde se guarda el código, sería mejor tener un arreglo
renglonCodigo=4
columnaCodigo=2
codigoADO = sheetADO.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPriceShoes = sheetPriceShoes.cell(row=renglonCodigo, column=columnaCodigo).value
codigoMabesa = sheetMabesa.cell(row=renglonCodigo, column=columnaCodigo).value
codigoMetaloides = sheetMetaloides.cell(row=renglonCodigo, column=columnaCodigo).value
codigoZuum = sheetZuum.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPoderJudicial = sheetPoderJudicial.cell(row=renglonCodigo, column=columnaCodigo).value
codigoSolutia = sheetSolutia.cell(row=renglonCodigo, column=columnaCodigo).value
codigoSumitomoTlaxcala = sheetSumitomoTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value
codigoAdientTlaxcala = sheetAdientTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value
codigoGestampFinsa = sheetGestampFinsa.cell(row=renglonCodigo, column=columnaCodigo).value
codigoAdientFinsa = sheetGestampFinsa.cell(row=renglonCodigo, column=columnaCodigo).value
codigoBrose = sheetBrose.cell(row=renglonCodigo, column=columnaCodigo).value
codigoKayser = sheetKayser.cell(row=renglonCodigo, column=columnaCodigo).value
codigoSmpFinsa = sheetSmpFinsa.cell(row=renglonCodigo, column=columnaCodigo).value
codigoROA = sheetROA.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPistones = sheetPistones.cell(row=renglonCodigo, column=columnaCodigo).value
codigoGestamp2 = sheetGestamp2.cell(row=renglonCodigo, column=columnaCodigo).value
codigoThyssen1 = sheetThyssen1.cell(row=renglonCodigo, column=columnaCodigo).value
codigoThyssen2 = sheetThyssen2.cell(row=renglonCodigo, column=columnaCodigo).value
codigoMaquinsa = sheetMaquinsa.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPosco = sheetPosco.cell(row=renglonCodigo, column=columnaCodigo).value
codigoBodega = sheetBodega.cell(row=renglonCodigo, column=columnaCodigo).value
codigoOficinas = sheetOficina.cell(row=renglonCodigo, column=columnaCodigo).value
codigoSumitomoAtlixco = sheetSumitomoAtlixco.cell(row=renglonCodigo, column=columnaCodigo).value
codigoFaurecia = sheetFaurecia.cell(row=renglonCodigo, column=columnaCodigo).value
codigoCojinetes = sheetCojinetes.cell(row=renglonCodigo, column=columnaCodigo).value
codigoAudi = sheetAudi.cell(row=renglonCodigo, column=columnaCodigo).value
codigoSmpTlaxcala = sheetSmpTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPjTlaxcala = sheetPjTlaxcala.cell(row=renglonCodigo, column=columnaCodigo).value
codigoInteva = sheetInteva.cell(row=renglonCodigo, column=columnaCodigo).value
codigoPakar = sheetPakar.cell(row=renglonCodigo, column=columnaCodigo).value

#Contador de lineas en hoja datos empleados
numeroEmpleados = sheetDatosEmpleados.max_row

#Lista de Empleados de control asistencias
listaEmpleados = []


# Recorrer la hoja de Datos Empleado y guardarlos en un arreglo
def leerDatosEmpleados():
    print("Leyendo Datos Empleados")
    for i in range(2, numeroEmpleados + 1):
        idEmpleado = sheetDatosEmpleados.cell(row=i, column=1).value
        nombreEmpleado = sheetDatosEmpleados.cell(row=i, column=2).value
        codigoComedor = sheetDatosEmpleados.cell(row=i, column=3).value
        nombreComedor = sheetDatosEmpleados.cell(row=i, column=4).value
        empleado = DatosEmpleado(idEmpleado, nombreEmpleado, codigoComedor, nombreComedor)
        listaEmpleados.append(empleado)



#Recorre la lista de empleados, la compara con un numero de comedor, y lo escribe a una hoja
def copiarEmpleadosComedor(numeroComedor,hojaComedor,listaEmpleados):
    sizeListaEmpleados = len(listaEmpleados)
    listaEmpleadosTMP = []
    for i in range(0,sizeListaEmpleados):
        comedorEmpleado = listaEmpleados[i].codigoComedor
        if numeroComedor == comedorEmpleado:
            listaEmpleadosTMP.append(listaEmpleados[i].idEmpleado)
    numeroEmpleadosTMP = len(listaEmpleadosTMP)
    for j in range(numeroEmpleadosTMP):
        renglonInicio = j+8
        hojaComedor.cell(row=renglonInicio,column=1).value = listaEmpleadosTMP[j]


#Zona de ejecución
def ejecutarAsignarEmpleadosAComedores():
    print("Ejecutando Asignando empleados.....")

    leerDatosEmpleados()  # Lee y sube a memoria todos los empleados
    # Repetir este código para cada hoja.
    copiarEmpleadosComedor(codigoADO, sheetADO, listaEmpleados)
    copiarEmpleadosComedor(codigoPriceShoes, sheetPriceShoes, listaEmpleados)
    copiarEmpleadosComedor(codigoMabesa, sheetMabesa, listaEmpleados)
    copiarEmpleadosComedor(codigoMetaloides, sheetMetaloides, listaEmpleados)
    copiarEmpleadosComedor(codigoZuum, sheetZuum, listaEmpleados)
    copiarEmpleadosComedor(codigoPoderJudicial, sheetPoderJudicial, listaEmpleados)
    copiarEmpleadosComedor(codigoSolutia, sheetSolutia, listaEmpleados)
    copiarEmpleadosComedor(codigoSumitomoTlaxcala, sheetSumitomoTlaxcala, listaEmpleados)
    copiarEmpleadosComedor(codigoAdientTlaxcala, sheetAdientTlaxcala, listaEmpleados)
    copiarEmpleadosComedor(codigoGestampFinsa, sheetGestampFinsa, listaEmpleados)
    copiarEmpleadosComedor(codigoAdientFinsa, sheetAdientFinsa, listaEmpleados)
    copiarEmpleadosComedor(codigoBrose, sheetBrose, listaEmpleados)
    copiarEmpleadosComedor(codigoKayser, sheetKayser, listaEmpleados)
    copiarEmpleadosComedor(codigoSmpFinsa, sheetSmpFinsa, listaEmpleados)
    copiarEmpleadosComedor(codigoROA, sheetROA, listaEmpleados)
    copiarEmpleadosComedor(codigoPistones, sheetPistones, listaEmpleados)
    copiarEmpleadosComedor(codigoGestamp2, sheetGestamp2, listaEmpleados)
    copiarEmpleadosComedor(codigoThyssen1, sheetThyssen1, listaEmpleados)
    copiarEmpleadosComedor(codigoThyssen2, sheetThyssen2, listaEmpleados)
    copiarEmpleadosComedor(codigoMaquinsa, sheetMaquinsa, listaEmpleados)
    copiarEmpleadosComedor(codigoPosco, sheetPosco, listaEmpleados)
    copiarEmpleadosComedor(codigoBodega, sheetBodega, listaEmpleados)
    copiarEmpleadosComedor(codigoOficinas, sheetOficina, listaEmpleados)
    copiarEmpleadosComedor(codigoSumitomoAtlixco, sheetSumitomoAtlixco, listaEmpleados)
    copiarEmpleadosComedor(codigoFaurecia, sheetFaurecia, listaEmpleados)
    copiarEmpleadosComedor(codigoCojinetes, sheetCojinetes, listaEmpleados)
    copiarEmpleadosComedor(codigoAudi, sheetAudi, listaEmpleados)
    copiarEmpleadosComedor(codigoSmpTlaxcala, sheetSmpTlaxcala, listaEmpleados)
    copiarEmpleadosComedor(codigoPjTlaxcala, sheetPjTlaxcala, listaEmpleados)
    copiarEmpleadosComedor(codigoInteva, sheetInteva, listaEmpleados)
    copiarEmpleadosComedor(codigoPakar, sheetPakar, listaEmpleados)

    wb.save('ControlAsistenciaProcomin_V3.xlsx')  # guardar un archivo nuevo


ejecutarAsignarEmpleadosAComedores()